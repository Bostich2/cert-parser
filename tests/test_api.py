from __future__ import annotations

import json
from pathlib import Path

import httpx
import pytest
import respx
from fastapi.testclient import TestClient

from sert_parser.api.app import create_app
from sert_parser.config import get_settings
from sert_parser.version import get_version

ODATA_URL = "https://tech.eaeunion.org/odata/ConformityDocDetailsType"
AM_EXAMPLE = "ЕАЭС AM C-CN.АБ12.В.00001/24"


def test_index_page(client: TestClient) -> None:
    version = get_version()
    response = client.get("/")
    assert response.status_code == 200
    assert "Парсер сертификатов" in response.text
    assert f'meta name="app-version" content="{version}"' in response.text
    assert f"/static/app.js?v={version}" in response.text
    assert 'class="app-version"' in response.text
    assert "settings-toggle" in response.text
    assert "Сбросить кэш" not in response.text
    assert "Полная перезагрузка" not in response.text
    assert "/static/flags/kz.svg" in response.text
    assert "<h2>Из буфера</h2>" in response.text
    assert "<h2>Из Excel</h2>" in response.text
    assert "<h2>Из PDF</h2>" in response.text
    assert 'class="sources-grid"' in response.text
    assert "pdf-drop" in response.text
    assert "Выбрать файлы" in response.text
    assert "выбрать папку" in response.text
    assert "Записей на страницу" in response.text
    assert "Экспорт в Excel" in response.text
    assert 'class="file-label">Файлы</span>' not in response.text
    assert "Действует с" in response.text
    assert "<h2>Один номер</h2>" not in response.text
    assert "<h2>Список номеров</h2>" not in response.text


def test_clear_cache(admin_client: TestClient) -> None:
    response = admin_client.post("/api/cache/clear")
    assert response.status_code == 200
    body = response.json()
    assert "deleted" in body
    assert "message" in body
    assert "Кэш очищен" in body["message"]


def test_reload_service_restarts_runtime(admin_client: TestClient) -> None:
    version = get_version()
    first = admin_client.post("/api/reload")
    assert first.status_code == 200
    body = first.json()
    assert body["version"] == version
    assert body["generation"] >= 2
    assert "перезапущен" in body["message"].lower()
    assert version in body["message"]
    second = admin_client.post("/api/reload")
    assert second.status_code == 200
    assert second.json()["generation"] == body["generation"] + 1
    lookup = admin_client.post("/api/lookup", json={"numbers": ["просто текст"]})
    assert lookup.status_code == 200
    assert lookup.json()["results"][0]["error_code"] == "invalid_number"


def test_health_includes_version(admin_client: TestClient) -> None:
    response = admin_client.get("/health")
    assert response.status_code == 200
    body = response.json()
    assert body["version"] == get_version()
    assert body["generation"] >= 1
    assert "status" in body


def test_lookup_rejects_empty_batch(client: TestClient) -> None:
    response = client.post("/api/lookup", json={"numbers": ["  ", ""]})
    assert response.status_code == 400


def test_lookup_invalid_number(client: TestClient) -> None:
    response = client.post("/api/lookup", json={"numbers": ["просто текст"]})
    assert response.status_code == 200
    result = response.json()["results"][0]
    assert result["error_code"] == "invalid_number"
    assert result["trace"]
    assert any("не разобран" in step.lower() for step in result["trace"])


def test_lookup_armenia(client: TestClient) -> None:
    payload = {
        "value": [
            {
                "docId": AM_EXAMPLE,
                "docStartDate": "2024-01-15T03:00:00+03:00",
                "docValidityDate": "2029-01-14T03:00:00+03:00",
                "Id": "am-test-id-00001",
                "unifiedCountryCode": {"value": "AM"},
                "docStatusDetails": {"docStatusCode": "01"},
            }
        ]
    }
    with respx.mock:
        respx.get(ODATA_URL).mock(return_value=httpx.Response(200, json=payload))
        response = client.post("/api/lookup", json={"numbers": [AM_EXAMPLE]})
    assert response.status_code == 200
    result = response.json()["results"][0]
    assert result["error_code"] is None
    assert result["country_code"] == "AM"
    assert result["registry_id"] == "am-test-id-00001"
    assert result["url"].endswith("/am-test-id-00001")
    assert result["trace"]


def test_lookup_pdf_extracts_number_and_looks_up(client: TestClient) -> None:
    import fitz

    document = fitz.open()
    page = document.new_page()
    page.insert_htmlbox(
        fitz.Rect(50, 50, 550, 250),
        "<p>Сертификат ЕАЭС AM C-CN.АБ12.В.00001/24</p>",
    )
    pdf_bytes = document.tobytes()
    document.close()
    odata_payload = {
        "value": [
            {
                "docId": AM_EXAMPLE,
                "docStartDate": "2024-01-15T03:00:00+03:00",
                "docValidityDate": "2029-01-14T03:00:00+03:00",
                "Id": "am-test-id-00001",
                "unifiedCountryCode": {"value": "AM"},
                "docStatusDetails": {"docStatusCode": "01"},
            }
        ]
    }
    with respx.mock:
        respx.get(ODATA_URL).mock(return_value=httpx.Response(200, json=odata_payload))
        response = client.post(
            "/api/lookup-pdf",
            files={"file": ("cert.pdf", pdf_bytes, "application/pdf")},
        )
    assert response.status_code == 200
    body = response.json()
    assert body["extracted_numbers"]
    assert body["results"][0]["error_code"] is None
    assert body["results"][0]["country_code"] == "AM"


def test_extract_pdf_reads_number(client: TestClient) -> None:
    import fitz

    document = fitz.open()
    page = document.new_page()
    page.insert_htmlbox(
        fitz.Rect(50, 50, 550, 250),
        "<p>Сертификат ЕАЭС AM C-CN.АБ12.В.00001/24</p>",
    )
    pdf_bytes = document.tobytes()
    document.close()
    response = client.post(
        "/api/extract-pdf",
        files={"file": ("cert.pdf", pdf_bytes, "application/pdf")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["error_code"] is None
    assert body["numbers"] == [AM_EXAMPLE]
    assert body["extract_trace"]


def _ndjson_events(response) -> list[dict]:
    return [json.loads(line) for line in response.text.splitlines() if line.strip()]


def test_extract_pdf_stream_emits_steps(client: TestClient) -> None:
    import fitz

    document = fitz.open()
    page = document.new_page()
    page.insert_htmlbox(
        fitz.Rect(50, 50, 550, 250),
        "<p>Сертификат ЕАЭС AM C-CN.АБ12.В.00001/24</p>",
    )
    pdf_bytes = document.tobytes()
    document.close()
    response = client.post(
        "/api/extract-pdf/stream",
        files={"file": ("cert.pdf", pdf_bytes, "application/pdf")},
    )
    assert response.status_code == 200
    events = _ndjson_events(response)
    assert any(item["type"] == "step" for item in events)
    assert events[-1]["type"] == "done"
    assert events[-1]["numbers"] == [AM_EXAMPLE]


def test_extract_pdf_stream_returns_error_envelope(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    def boom(_payload, _settings):
        raise RuntimeError("OCR exploded")

    monkeypatch.setattr("sert_parser.application.extract_service.extract_numbers_from_pdf", boom)
    response = client.post(
        "/api/extract-pdf/stream",
        files={"file": ("cert.pdf", b"%PDF-1.4", "application/pdf")},
    )
    assert response.status_code == 200
    events = _ndjson_events(response)
    assert events[-1]["type"] == "error"
    assert events[-1]["detail"] == "Внутренняя ошибка обработки"


def test_lookup_stream_emits_steps(client: TestClient) -> None:
    response = client.post("/api/lookup/stream", json={"numbers": ["просто текст"]})
    assert response.status_code == 200
    events = _ndjson_events(response)
    assert any(item.get("text") and "не разобран" in item["text"].lower() for item in events)
    assert events[-1]["type"] == "done"
    assert events[-1]["result"]["error_code"] == "invalid_number"


def test_extract_pdf_without_numbers(client: TestClient) -> None:
    import fitz

    document = fitz.open()
    document.new_page()
    payload = document.tobytes()
    document.close()
    response = client.post(
        "/api/extract-pdf",
        files={"file": ("empty.pdf", payload, "application/pdf")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["error_code"] == "no_numbers_in_pdf"
    assert body["numbers"] == []
    assert "extract_trace" in body


def test_lookup_pdf_without_numbers(client: TestClient) -> None:
    import fitz

    document = fitz.open()
    document.new_page()
    payload = document.tobytes()
    document.close()
    response = client.post(
        "/api/lookup-pdf",
        files={"file": ("empty.pdf", payload, "application/pdf")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["error_code"] == "no_numbers_in_pdf"
    assert body["results"] == []
    assert "extract_trace" in body


def test_extract_xlsx_reads_column_a(client: TestClient) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    example = "ЕАЭС BY/112 02.01. ТР018 010.02 00276"
    second = "ЕАЭС RU С-CN.СБ21.А.00039/19"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = "Номер"
    sheet["A2"] = example
    sheet["A3"] = ""
    sheet["A4"] = second
    buffer = BytesIO()
    workbook.save(buffer)
    response = client.post(
        "/api/extract-xlsx",
        files={
            "file": (
                "numbers.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["error_code"] is None
    assert body["numbers"] == [example, second]


def test_extract_xlsx_without_numbers(client: TestClient) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = "Номер"
    buffer = BytesIO()
    workbook.save(buffer)
    response = client.post(
        "/api/extract-xlsx",
        files={
            "file": (
                "empty.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["error_code"] == "no_numbers_in_xlsx"
    assert body["numbers"] == []


def test_extract_xlsx_rejects_invalid_file(client: TestClient) -> None:
    response = client.post(
        "/api/extract-xlsx",
        files={"file": ("broken.xlsx", b"not-excel", "application/octet-stream")},
    )
    assert response.status_code == 400


def test_export_xlsx_returns_attachment(client: TestClient) -> None:
    response = client.post(
        "/api/export-xlsx",
        json={
            "results": [
                {
                    "query": "ЕАЭС BY/112 02.01. ТР018 010.02 00276",
                    "official_number": "ЕАЭС BY/112 02.01. ТР018 010.02 00276",
                    "country_code": "BY",
                    "status": "Действует",
                    "status_code": "01",
                    "cached": True,
                }
            ]
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=\"sert-parser-results.xlsx\"" in response.headers["content-disposition"]
    assert response.content.startswith(b"PK")


def test_extract_pdf_reports_truncation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("CACHE_PATH", str(tmp_path / "cache.sqlite"))
    monkeypatch.setenv("LOOKUP_DELAY_SECONDS", "0")
    monkeypatch.setenv("MAX_BATCH_SIZE", "2")
    monkeypatch.setenv("PDF_OCR_ENABLED", "false")
    monkeypatch.setenv("RATE_LIMIT_UPLOAD", "1000/minute")
    get_settings.cache_clear()
    app = create_app()
    numbers = [
        "ЕАЭС BY/112 02.01. ТР018 010.02 00276",
        "ЕАЭС BY/112 02.01. ТР018 010.02 00277",
        "ЕАЭС BY/112 02.01. ТР018 010.02 00278",
    ]
    monkeypatch.setattr(
        "sert_parser.application.extract_service.extract_numbers_from_pdf",
        lambda _payload, _settings: list(numbers),
    )
    with TestClient(app) as test_client:
        response = test_client.post(
            "/api/extract-pdf",
            files={"file": ("many.pdf", b"%PDF-1.4", "application/pdf")},
        )
    get_settings.cache_clear()
    assert response.status_code == 200
    body = response.json()
    assert body["truncated"] is True
    assert body["total_found"] == 3
    assert body["numbers"] == numbers[:2]
    assert "3" in body["warning"]


def test_extract_xlsx_rejects_too_many_numbers(client: TestClient) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for index in range(101):
        sheet.cell(row=index + 1, column=1, value=f"ЕАЭС BY/112 02.01. ТР018 010.02 {index:05d}")
    buffer = BytesIO()
    workbook.save(buffer)
    response = client.post(
        "/api/extract-xlsx",
        files={
            "file": (
                "many.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 400
    assert "Слишком много номеров" in response.json()["detail"]


def test_extract_xlsx_rejects_oversized_file(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("XLSX_MAX_BYTES", "10")
    monkeypatch.setenv("RATE_LIMIT_UPLOAD", "1000/minute")
    get_settings.cache_clear()
    app = create_app()
    with TestClient(app) as test_client:
        response = test_client.post(
            "/api/extract-xlsx",
            files={"file": ("big.xlsx", b"x" * 20, "application/octet-stream")},
        )
    get_settings.cache_clear()
    assert response.status_code == 413
    assert "больше" in response.json()["detail"]


def test_lookup_during_reload_returns_503(client: TestClient) -> None:
    client.app.state.reload_in_progress = True
    try:
        response = client.post("/api/lookup", json={"numbers": ["просто текст"]})
    finally:
        client.app.state.reload_in_progress = False
    assert response.status_code == 503
    assert "перезапускается" in response.json()["detail"]


def test_reload_waits_for_in_flight_lookup(admin_client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    import asyncio
    import threading
    import time

    from sert_parser.application.lookup_service import LookupService

    original_lookup_one = LookupService.lookup_one

    async def slow_lookup_one(self, raw: str):
        await asyncio.sleep(0.35)
        return await original_lookup_one(self, raw)

    monkeypatch.setattr(LookupService, "lookup_one", slow_lookup_one)

    lookup_result: dict = {}

    def run_lookup():
        lookup_result["response"] = admin_client.post(
            "/api/lookup",
            json={"numbers": ["просто текст"]},
        )

    thread = threading.Thread(target=run_lookup)
    thread.start()
    time.sleep(0.05)
    reload_response = admin_client.post("/api/reload")
    thread.join(timeout=5)
    assert reload_response.status_code == 200
    assert lookup_result["response"].status_code == 200
    assert lookup_result["response"].json()["results"][0]["error_code"] == "invalid_number"


def test_clear_cache_during_reload_returns_503(admin_client: TestClient) -> None:
    admin_client.app.state.reload_in_progress = True
    try:
        response = admin_client.post("/api/cache/clear")
    finally:
        admin_client.app.state.reload_in_progress = False
    assert response.status_code == 503


def test_reload_blocked_while_lookup_active(
    admin_client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import asyncio
    import importlib
    import threading
    import time

    from sert_parser.application.lookup_service import LookupService

    api_module = importlib.import_module("sert_parser.api.app")
    monkeypatch.setattr(api_module, "RELOAD_WAIT_TIMEOUT_SECONDS", 0.05)

    original_lookup_one = LookupService.lookup_one

    async def slow_lookup_one(self, raw: str):
        await asyncio.sleep(0.4)
        return await original_lookup_one(self, raw)

    monkeypatch.setattr(LookupService, "lookup_one", slow_lookup_one)

    lookup_thread = threading.Thread(
        target=lambda: admin_client.post("/api/lookup", json={"numbers": ["просто текст"]}),
    )
    lookup_thread.start()
    time.sleep(0.05)
    reload_response = admin_client.post("/api/reload")
    lookup_thread.join(timeout=5)
    assert reload_response.status_code == 409
    assert "активны запросы" in reload_response.json()["detail"]


def test_stream_lookup_reserves_slot_before_work(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import importlib
    import threading
    import time

    api_module = importlib.import_module("sert_parser.api.app")
    gate = threading.Event()

    original_stream = api_module.stream_async_work

    async def gated_stream(work):
        gate.wait(timeout=2)
        async for chunk in original_stream(work):
            yield chunk

    monkeypatch.setattr(api_module, "stream_async_work", gated_stream)
    stream_thread = threading.Thread(
        target=lambda: client.post(
            "/api/lookup/stream",
            json={"numbers": ["просто текст"]},
        ),
    )
    stream_thread.start()
    time.sleep(0.05)
    try:
        assert int(client.app.state.active_lookups) >= 1
    finally:
        gate.set()
        stream_thread.join(timeout=5)
        client.app.state.active_lookups = 0
