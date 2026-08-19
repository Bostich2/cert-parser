from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from sert_parser.domain.errors import XlsxReadError
from sert_parser.infrastructure.xlsx import build_results_xlsx, extract_numbers_from_xlsx

EXAMPLE = "ЕАЭС BY/112 02.01. ТР018 010.02 00276"
SECOND = "ЕАЭС RU С-CN.СБ21.А.00039/19"


def xlsx_bytes(rows: list[object], extra_column: object | None = "ignored") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for index, value in enumerate(rows, start=1):
        sheet.cell(row=index, column=1, value=value)
        if extra_column is not None:
            sheet.cell(row=index, column=2, value=extra_column)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_extracts_column_a_and_skips_empty() -> None:
    payload = xlsx_bytes([EXAMPLE, "", None, SECOND, "  "])
    assert extract_numbers_from_xlsx(payload) == [EXAMPLE, SECOND]


def test_skips_header_row() -> None:
    payload = xlsx_bytes(["Номер сертификата", EXAMPLE, SECOND])
    assert extract_numbers_from_xlsx(payload) == [EXAMPLE, SECOND]


def test_keeps_first_row_when_not_header() -> None:
    payload = xlsx_bytes([EXAMPLE, SECOND])
    assert extract_numbers_from_xlsx(payload) == [EXAMPLE, SECOND]


def test_empty_file_raises() -> None:
    with pytest.raises(XlsxReadError, match="пустой"):
        extract_numbers_from_xlsx(b"")


def test_invalid_xlsx_raises() -> None:
    with pytest.raises(XlsxReadError, match="открыть"):
        extract_numbers_from_xlsx(b"not-an-excel-file")


def test_too_many_numbers_raises() -> None:
    rows = [f"ЕАЭС BY/112 02.01. ТР018 010.02 0027{i}" for i in range(3)]
    payload = xlsx_bytes(rows, extra_column=None)
    with pytest.raises(XlsxReadError, match="Слишком много номеров"):
        extract_numbers_from_xlsx(payload, max_batch_size=2)


def test_build_results_xlsx_writes_headers_and_rows() -> None:
    payload = build_results_xlsx(
        [
            {
                "query": EXAMPLE,
                "official_number": EXAMPLE,
                "country_code": "BY",
                "url": "https://example.test/card",
                "valid_from": "2024-01-01",
                "valid_until": "2025-01-01",
                "status": "Действует",
                "status_code": "01",
                "error": "",
                "error_code": "",
                "cached": True,
            }
        ]
    )
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        assert sheet is not None
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0] == (
            "Запрос",
            "Номер",
            "Страна",
            "Ссылка",
            "Действует с",
            "Действует до",
            "Статус",
            "Код статуса",
            "Ошибка",
            "Код ошибки",
            "Из кэша",
        )
        assert rows[1] == (
            EXAMPLE,
            EXAMPLE,
            "BY",
            "https://example.test/card",
            "2024-01-01",
            "2025-01-01",
            "Действует",
            "01",
            None,
            None,
            "Да",
        )
    finally:
        workbook.close()
