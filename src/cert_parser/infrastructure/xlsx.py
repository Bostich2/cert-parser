from __future__ import annotations

from io import BytesIO

from cert_parser.domain.errors import XlsxReadError

_HEADER_HINTS = (
    "номер",
    "number",
    "сертификат",
    "наименован",
    "товар",
    "продукц",
    "product",
)


def extract_numbers_from_xlsx(payload: bytes, *, max_batch_size: int | None = None) -> list[str]:
    if not payload:
        raise XlsxReadError("Файл Excel пустой")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise XlsxReadError("Не установлен модуль openpyxl") from exc
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise XlsxReadError("Не удалось открыть Excel") from exc
    try:
        sheet = workbook.active
        if sheet is None:
            raise XlsxReadError("В книге Excel нет листа")
        numbers: list[str] = []
        for index, row in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True)):
            cell = row[0] if row else None
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            if index == 0 and _looks_like_header(text):
                continue
            numbers.append(text)
        if max_batch_size is not None and len(numbers) > max_batch_size:
            raise XlsxReadError(
                f"Слишком много номеров в Excel. Максимум {max_batch_size}"
            )
        return numbers
    finally:
        workbook.close()


_DOC_KIND_LABELS = {
    "certificate": "Сертификат",
    "declaration": "Декларация",
}


def build_results_xlsx(rows: list[dict]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise XlsxReadError("Не установлен модуль openpyxl") from exc

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise XlsxReadError("В книге Excel нет листа")
    sheet.title = "Результаты"
    include_product = any(row.get("product_name") or row.get("doc_kind") for row in rows)
    headers = ["Запрос"]
    if include_product:
        headers.extend(["Продукция", "Вид"])
    headers.extend(
        [
            "Номер",
            "Страна",
            "Ссылка",
            "PDF",
            "Действует с",
            "Действует до",
            "Статус",
            "Код статуса",
            "Ошибка",
            "Код ошибки",
            "Из кэша",
        ]
    )
    sheet.append(headers)
    for item in rows:
        row = [item.get("query") or ""]
        if include_product:
            kind = item.get("doc_kind") or ""
            row.extend(
                [
                    item.get("product_name") or "",
                    _DOC_KIND_LABELS.get(kind, kind),
                ]
            )
        row.extend(
            [
                item.get("official_number") or item.get("normalized") or item.get("query") or "",
                item.get("country_code") or "",
                item.get("url") or "",
                item.get("pdf_url") or "",
                item.get("valid_from") or "",
                item.get("valid_until") or "",
                item.get("status") or "",
                item.get("status_code") or "",
                item.get("error") or "",
                item.get("error_code") or "",
                "Да" if item.get("cached") else "Нет",
            ]
        )
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _looks_like_header(text: str) -> bool:
    lowered = text.lower()
    return any(hint in lowered for hint in _HEADER_HINTS)
